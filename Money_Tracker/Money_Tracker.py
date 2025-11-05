import datetime
from openpyxl import Workbook,load_workbook
import os
import json
import pandas as pd
import sys
from pathlib import Path

current_dir = os.path.dirname(os.path.abspath(__file__))
BASE = Path(current_dir)

class Details:
    def write_in_excel(self,date,day,time,amount,type_of_transaction,category,account,notes=None):
        data=[date,day,time,amount,type_of_transaction,category,account,notes]
        headers=["Date","Day","Time","Amount","Type","Category","Account","Notes"]
        filename=os.path.join(current_dir,"Money Tracker.xlsx")
        if not os.path.exists(filename):
            wb=Workbook()
            ws=wb.active
            ws.append(headers)
        else:
            wb=load_workbook(filename)
            ws=wb.active
        ws.append(data)
        wb.save(filename)

    def get_date_time(self):
        date=datetime.date.today().strftime("%d/%m/%y")
        day=datetime.datetime.now().strftime("%A")
        time=datetime.datetime.now().strftime("%I:%M %p")
        return date,day,time

    def get_amount(self):
        while True:
            raw = input("Enter the transaction's amount: ").replace(',', '').replace('₹', '').strip()
            try:
                amount = float(raw)
                amount = int(amount) if amount.is_integer() else amount
                break
            except ValueError:
                print("Enter valid amount..")
        return amount


    def get_income_category(self):
        filename=os.path.join(current_dir,"Income_category.json")
        lines=0
        if not os.path.exists(filename):
            print("No categories exists yet! Add new income categories")
            lines+=1
            add_income_category(lines)
        while True:
            with open(filename,'r',encoding='utf-8') as f:
                data=json.load(f)
            categories=data.get("categories",{})
            categories_list=list(categories.keys())
            print(f"0.Add new categories")
            lines+=1
            for i in range(1,len(categories_list)+1):
                print(f"{i}.{categories_list[i-1]}")
            
            while True:
                cat_number=int(input("Enter the corresponding category number:"))
                if cat_number == 0:
                    lines+=len(categories_list)+1
                    add_income_category(lines)
                    break
                else:
                    if 1<=cat_number<=len(categories_list):
                        selected_cat=categories_list[cat_number-1]
                        subcategories=list(categories[selected_cat])
                        if not subcategories:
                            return categories_list[cat_number-1]
                        else:
                            for i in range(1,len(subcategories)+1):
                                print(f"{i}.{subcategories[i-1]}")
                            while True:
                                subcat_number=int(input("Enter the corresponding sub-category number:"))
                                if 1<=subcat_number<=len(subcategories):
                                    return_value=f"{categories_list[cat_number-1]} : {(subcategories[subcat_number-1])}"
                                    return return_value
                                else:
                                    print(f"Enter valid number(1,{len(subcategories)})")
                                    continue
                    else:
                        print(f"Enter valid number(1,{len(categories_list)})")
                        continue

    def get_expense_category(self):
        filename=os.path.join(current_dir,"Expense_category.json")
        lines=0
        if not os.path.exists(filename):
            print("No categories exists yet! Add new expense categories")
            lines+=1
            add_expense_category(lines)
        while True:     
            with open(filename,'r',encoding='utf-8') as f:
                data=json.load(f)
            categories=data.get("categories",{})
            categories_list=list(categories.keys())
            print(f"0.Add new categories")
            lines+=1
            for i in range(1,len(categories_list)+1):
                print(f"{i}.{categories_list[i-1]}")
            while True:
                cat_number=int(input("Enter the corresponding category number:"))
                if cat_number == 0:
                    lines+=len(categories_list)+1
                    add_expense_category(lines)
                    break
                else:
                    if 1<=cat_number<=len(categories_list):
                        selected_cat=categories_list[cat_number-1]
                        subcategories=list(categories[selected_cat])
                        if not subcategories:
                            return categories_list[cat_number-1]
                        else:
                            for i in range(1,len(subcategories)+1):
                                print(f"{i}.{subcategories[i-1]}")
                            while True:
                                subcat_number=int(input("Enter the corresponding sub-category number:"))
                                if 1<=subcat_number<=len(subcategories):
                                    return_value=f"{categories_list[cat_number-1]} : {(subcategories[subcat_number-1])}"
                                    return return_value
                                else:
                                    print(f"Enter valid number(1,{len(subcategories)})")
                                    continue
                    else:
                        print(f"Enter valid number(1,{len(categories_list)})")
                        continue

    def get_account(self):
        filename=os.path.join(current_dir,"Accounts.json")
        lines=0
        
        if not os.path.exists(filename):
            print("No accounts found! Add new accounts")
            lines+=1
            add_account(lines)
        while True:
            with open(filename,'r',encoding='utf-8') as f:
                data=json.load(f)
            accounts=data.get("accounts",{})
            type_account_list=list(accounts.keys())
            print(f"0.Add New Accounts")
            lines+=1
            for i in range(1,len(type_account_list)+1):
                print(f"{i}.{type_account_list[i-1]}")
            while True:
                selected_type_index=int(input("Enter the corresponding choice:"))
                if selected_type_index==0:
                    lines+=len(type_account_list)+1
                    add_account(lines)
                    break
                else:
                    if 1<=selected_type_index<=len(type_account_list):
                        selected_type=type_account_list[selected_type_index-1]
                        account_name_list=list(accounts[selected_type])
                        if not account_name_list:
                            print("No accounts under this type yet. Please add accounts.")
                            return selected_type, None
                        else:
                            for i in range(1,len(account_name_list)+1):
                                print(f"{i}.{account_name_list[i-1]}")
                            while True:
                                selected_account_index=int(input("Enter the corresponding account name's index (or exit to exit the option):"))
                                if 1<=selected_account_index<=len(account_name_list):
                                    return type_account_list[selected_type_index-1],account_name_list[selected_account_index-1]
                                else:
                                    print(f"Enter valid number(1,{len(account_name_list)})")
                                    continue
                    else:
                        print(f"Enter valid number(1,{len(type_account_list)})")
                        continue
            return None,None
    def get_notes(self):
        notes=input("Wanna have any note for this transaction:").title()
        return notes
    
    def get_transfer_accounts(self):
        filename = p("Accounts.json")
        data = safe_load_json(filename, {"accounts": {}})
        accounts = data.get("accounts", {})

        count, acc_map = 1, {}
        for type_acc, acc_name in accounts.items():
            print(f"{type_acc}:")
            for acc, bal in acc_name.items():
                print(f"  {count}. {acc} : {bal}")
                acc_map[count] = (type_acc, acc)
                count += 1

        max_idx = len(acc_map)
        if max_idx < 2:
            print("Need at least two accounts to transfer.")
            return None

        while True:
            opt_from = input(f"\nEnter FROM index (or 'q' to cancel): ").strip().lower()
            if opt_from == 'q':
                return None
            opt_to = input(f"Enter TO index (or 'q' to cancel): ").strip().lower()
            if opt_to == 'q':
                return None
            try:
                from_acc = int(opt_from)
                to_acc = int(opt_to)
                if not (1 <= from_acc <= max_idx and 1 <= to_acc <= max_idx):
                    print(f"Enter valid account choice (1-{max_idx})")
                    continue
                type_from, acc_from = acc_map[from_acc]
                type_to, acc_to = acc_map[to_acc]
                if (type_from, acc_from) == (type_to, acc_to):
                    print("Source and destination cannot be the same.")
                    continue
                return type_from, acc_from, type_to, acc_to
            except ValueError:
                print("Enter valid integer.")

    def get_and_transfer_balances(self,amount,type_from,acc_from,type_to,acc_to):
        filename=os.path.join(current_dir,"Accounts.json")
        with open(filename,'r',encoding='utf-8') as f:
            data=json.load(f)
        accounts=data.get("accounts",{})
        from_bal = accounts[type_from][acc_from]
        to_bal = accounts[type_to][acc_to]
        if from_bal<amount:
            return False
        else:
            accounts[type_from][acc_from]-=amount
            accounts[type_to][acc_to]+=amount

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump({"accounts": accounts}, f, indent=4)
        return True

def p(*parts): 
    return BASE.joinpath(*parts)

def safe_load_json(path, default):
    path = Path(path)
    if not path.exists() or path.stat().st_size == 0:
        return default
    try:
        return json.loads(path.read_text(encoding='utf-8'))
    except json.JSONDecodeError:
        return default

def safe_dump_json(path, data):
    Path(path).write_text(json.dumps(data, indent=4), encoding='utf-8')

def add_income_category(lines):
    filename=os.path.join(current_dir,"Income_category.json")
    if os.path.exists(filename) and os.path.getsize(filename)>0:
        with open(filename,'r',encoding='utf-8') as f:
            data=json.load(f)
        categories=data.get("categories",{})
    else:
        categories={}

    while True:
        maincat=input("Enter the category name (or type 'n' to exit the option):").title()
        lines+=1
        if maincat.lower()=="n":
            break
        subcats=[]
        while True:
            subcat=input(f"Enter the subcategory for {maincat} (or type 'n' to exit the option) :").title()
            lines+=1
            if subcat.lower()=="n":
                break
            subcats.append(subcat)
        if maincat in categories:
            existing_subs=categories[maincat]
            for s in subcats:
                if s not in existing_subs:
                    existing_subs.append(s)
        else:
            categories[maincat]=subcats
    
    with open(filename,'w',encoding='utf-8') as f:
        json.dump({"categories":categories},f,indent=4)
    print("Categories added successfully..")
    lines+=1
    clear_lines(lines)

def add_expense_category(lines):
    filename=os.path.join(current_dir,"Expense_category.json")
    if os.path.exists(filename) and os.path.getsize(filename)>0:
        with open(filename,'r',encoding='utf-8') as f:
            data=json.load(f)
        categories=data.get("categories",{})
    else:
        categories={}

    while True:
        maincat=input("Enter the category name (or type 'n' to exit the option):").title()
        lines+=1
        if maincat.lower()=="n":
            break
        subcats=[]
        while True:
            subcat=input(f"Enter the subcategory for {maincat} (or type 'n' to exit the option) :").title()
            lines+=1
            if subcat.lower()=="n":
                break
            subcats.append(subcat)
        if maincat in categories:
            existing_subs=categories[maincat]
            for s in subcats:
                if s not in existing_subs:
                    existing_subs.append(s)
        else:
            categories[maincat]=subcats
    
    with open(filename,'w',encoding='utf-8') as f:
        json.dump({"categories":categories},f,indent=4)
    print("Categories added successfully..")
    lines+=1
    clear_lines(lines)

def add_account(lines):
    filename = p("Accounts.json")
    data = safe_load_json(filename, {"accounts": {}})
    accounts = data.get("accounts", {})
    
    # if os.path.exists(filename):
    #     with open(filename,'r',encoding='utf-8') as f:
    #         data=json.load(f)
    #         accounts=data.get("accounts",{})
    # else:
    #     accounts={}

    while True:
        type_account=input(f"Enter the type of account (or type 'n' to exit the option):").title()
        lines+=1
        if type_account.lower()=='n':
            break

        account_dict = {}
        while True:
            account_name=input(f"Enter the name of account for {type_account} (or type 'n' to exit the option):").title()
            lines+=1
            if account_name.lower()=='n':
                break
            while True:
                try:
                    raw = input(f"Add money for {account_name} (or 0): ").replace(',', '').replace('₹','').strip()
                    money = float(raw)
                    lines+=1
                    balance = int(money) if money.is_integer() else money
                    break
                except ValueError:
                    print("Please enter a valid number.")
                    lines+=1
            
            account_dict[account_name] = balance

        # Update accounts data
        if type_account in accounts:
            accounts[type_account].update(account_dict)
        else:
            accounts[type_account] = account_dict

        # Save the updated accounts data
    safe_dump_json(filename, {"accounts": accounts})
    clear_lines(lines)

def clear_lines(n):
    if os.name != 'nt':
        for _ in range(n):
            sys.stdout.write("\x1b[1A")
            sys.stdout.write("\x1b[2K")
        sys.stdout.flush()

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def update_account_balance(type_account, account_name, amount):
    filename = os.path.join(current_dir, "Accounts.json")
    with open(filename, "r", encoding="utf-8") as f:
        data = json.load(f)
    accounts = data.get("accounts", {})
    if type_account not in accounts or account_name not in accounts[type_account]:
        print("Account not found. Please check your Accounts.json.")
        return False
    current_balance = accounts[type_account][account_name]
    new_balance = current_balance + amount  # works for both income (+) and expense (-)
    if new_balance < 0:
        return False
    accounts[type_account][account_name] = new_balance
    with open(filename, "w", encoding="utf-8") as f:
        json.dump({"accounts": accounts}, f, indent=4)
    return True

def add_income():
    income = Details()
    date, day, time = income.get_date_time()
    amount = income.get_amount()
    type_of_transaction = "Income"
    category = income.get_income_category()
    type_acc, account = income.get_account()
    account_value=f"{type_acc} : {account}"
    notes = income.get_notes()

    update_account_balance(type_acc, account, amount)

    income.write_in_excel(date, day, time, amount, type_of_transaction, category, account_value, notes)

def add_expense():
    expense = Details()
    date, day, time = expense.get_date_time()
    amount = expense.get_amount()
    type_of_transaction = "Expense"
    category = expense.get_expense_category()
    while True:
        type_acc, account = expense.get_account()
        

        # Try to deduct balance
        success = update_account_balance(type_acc, account, -amount)

        if success:
            account_value=f"{type_acc} : {account}"
            notes = expense.get_notes()
            expense.write_in_excel(date, day, time, -amount, type_of_transaction, category, account_value, notes)
            break
        else:
            print("\n Insufficient Balance!!\n")
            retry = input("Would you like to choose another account? (y/n): ").lower()
            if retry != "y":
                print("Transaction cancelled.")
                break

def transfer_between_accounts():
    transfer=Details()
    date,day,time=transfer.get_date_time()
    amount=transfer.get_amount()
    type_of_transaction = "Transfer"
    category="Transfer between accounts"
    while True:
        type_from,acc_from,type_to,acc_to=transfer.get_transfer_accounts()
        
        success = transfer.get_and_transfer_balances(amount,type_from,acc_from,type_to,acc_to)

        if success:
            account_value=f"From {acc_from} - To {acc_to}"
            notes = transfer.get_notes()
            transfer.write_in_excel(date, day, time, amount, type_of_transaction, category, account_value, notes)
            break
        else:
            print("\n Insufficient Balance!!\n")
            retry = input("Would you like to choose another account? (y/n): ").lower()
            if retry != "y":
                print("Transaction cancelled.")
                break

def view_summary():

    excel_path = os.path.join(current_dir, "Money Tracker.xlsx")
    accounts_path = os.path.join(current_dir, "Accounts.json")
    if not os.path.exists(excel_path):
        print("No transaction data found yet.")
        return

    # Load Excel data
    df = pd.read_excel(excel_path)
    if df.empty:
        print("No transaction rows yet.")
        return

    print("\n==================== SUMMARY ====================\n")

    total_income = df.loc[df["Type"] == "Income", "Amount"].sum()
    total_expense = abs(df.loc[df["Type"] == "Expense", "Amount"].sum())
    net_savings = total_income - total_expense

    print("OVERALL SUMMARY")
    print(f" Total Income  = ₹{total_income:,.0f}")
    print(f" Total Expense = ₹{total_expense:,.0f}")
    print(f" Net Savings   = ₹{net_savings:,.0f}\n")
    print("-" * 55)

    if os.path.exists(accounts_path):
        with open(accounts_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        accounts=data.get("accounts", {})
        print("ACCOUNT-WISE BALANCES")
        for acc_type, acc in accounts.items():
            balances = ", ".join([f"{name}: ₹{bal:,}" for name, bal in acc.items()])
            print(f"{acc_type} → {balances}")
    else:
        print("No account data available.")
    print()
    print("-" * 55)

    expense_df = df[df["Type"] == "Expense"]
    if not expense_df.empty:
        print("CATEGORY-WISE SPENDING")
        category_summary = expense_df.groupby("Category")["Amount"].sum().abs().sort_values(ascending=False)
        for cat, val in category_summary.items():
            print(f"{cat}: ₹{val:,.0f}")
    else:
        print("No expense transactions found.")
    print("-" * 55)

    # 4️⃣ Monthly Summary (Income vs Expense per month)
    print("MONTHLY SUMMARY")
    df["Month-Year"] = pd.to_datetime(df["Date"], format="%d/%m/%y", errors='coerce').dt.strftime("%B %Y")
    monthly_summary = df.groupby(["Month-Year", "Type"])["Amount"].sum().unstack(fill_value=0)
    for month, row in monthly_summary.iterrows():
        inc = row.get("Income", 0)
        exp = abs(row.get("Expense", 0))
        print(f"📅 {month} → Income: ₹{inc:,.0f}, Expense: ₹{exp:,.0f}")
    print("-" * 55)

    # 5️⃣ Recent Transactions
    print("RECENT TRANSACTIONS (Last 5)\n")
    recent = df.tail(5)[["Date", "Type", "Amount", "Account", "Category"]]

    # Print header
    print(f"{'Date':<10} | {'Type':<10} | {'Amount':>9} | {'Account':<43} | {'Category'}")
    print("-" * 100)

    # Print rows neatly aligned
    for _, row in recent.iterrows():
        print(f"{row['Date']:<10} | {row['Type']:<10} | ₹{row['Amount']:>8,.0f} | {row['Account']:<43} | {row['Category']}")

    print("-" * 100)

def main():
    clear_screen()
    transaction_added=False
    while True:
        print("\n--------------------Welcome to Your Personal Money Tracker--------------------")
        print("\t1.Income")
        print("\t2.Expense")
        print("\t3.View Summary")
        print("\t4.Transfer between accounts")
        print("\t5.Add Income category")
        print("\t6.Add Expense category")
        print("\t7.Add accounts")
        print("\t8.Exit the tracker")
        
        try:
            type_of_task=int(input("\nEnter the type of transaction(1-8):"))
        except ValueError:
            clear_screen()
            print("Please enter integer only(1-8)!!")
            continue
        match(type_of_task):
            case 1:
                print()
                add_income()
                transaction_added=True
            case 2:
                print()
                add_expense()
                transaction_added=True
            case 3:
                clear_screen()
                print()
                view_summary()
                break
            case 4:
                print()
                transfer_between_accounts()
                transaction_added=True
            case 5:
                print()
                lines=0
                add_income_category(lines)
                transaction_added=True
            case 6:
                print()
                lines=0
                add_expense_category(lines)
                transaction_added=True
            case 7:
                print()
                lines=0
                add_account(lines)
                transaction_added=True
            case 8:
                clear_screen()
                print("Exiting the program !! Thank you")
                break
            case _:
                clear_screen()
                print("Enter a valid task number")
        if transaction_added:
            wanna_continue=input("Do you want to add another transaction(y,n):").lower()
            clear_screen()
            if wanna_continue=='n':
                print("Transactions logged successfully")
                print("Exiting the program !! Thank you")
                break


if __name__=="__main__":
    main()